"""
Excel file management module for SharePoint Excel Manager
Handles downloading, opening, and analyzing Excel files
"""
import logging
import os
import tempfile
from pathlib import Path
from typing import Dict, List, Optional

import openpyxl
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows

logger = logging.getLogger(__name__)


class ExcelManager:
    """Manages Excel file operations including downloading, opening, and table extraction"""
    
    def __init__(self, sharepoint_client):
        self.sharepoint_client = sharepoint_client
        self.temp_dir = None
        self.current_workbook = None
        self.current_file_path = None
        self.current_file_info = None
    
    def __enter__(self):
        """Context manager entry - create temp directory"""
        self.temp_dir = tempfile.mkdtemp(prefix="sharepoint_excel_")
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        """Context manager exit - cleanup temp files"""
        self.cleanup()
    
    def cleanup(self):
        """Clean up temporary files and close workbook"""
        if self.current_workbook:
            try:
                self.current_workbook.close()
            except:
                pass
            self.current_workbook = None
        
        if self.temp_dir and os.path.exists(self.temp_dir):
            try:
                import shutil
                shutil.rmtree(self.temp_dir)
            except Exception as e:
                logger.warning(f"Could not remove temp directory: {e}")
    
    async def download_and_open_excel_file(self, file_info: Dict) -> bool:
        """Download and open Excel file without displaying it"""
        try:
            logger.info(f"Downloading Excel file: {file_info['name']}")
            
            # Create temp file path
            file_name = file_info['name']
            self.current_file_path = os.path.join(self.temp_dir, file_name)
            self.current_file_info = file_info
            
            # Download the file
            success = await self.sharepoint_client.download_file(file_info, self.current_file_path)
            
            if not success:
                raise Exception("Failed to download file from SharePoint")
            
            # Open the Excel file (without displaying)
            logger.info(f"Opening Excel file: {file_name}")
            self.current_workbook = openpyxl.load_workbook(self.current_file_path, read_only=True)
            
            logger.info(f"Successfully opened Excel file: {file_name}")
            return True
            
        except Exception as e:
            logger.error(f"Error downloading/opening Excel file: {e}")
            return False
    
    def get_available_tables(self) -> List[Dict]:
        """Extract list of available tables/worksheets from the Excel file"""
        if not self.current_workbook:
            logger.error("No workbook is currently open")
            return []
        
        try:
            tables = []
            
            # Get all worksheets
            for sheet_name in self.current_workbook.sheetnames:
                worksheet = self.current_workbook[sheet_name]
                
                # Get sheet dimensions
                max_row = worksheet.max_row
                max_col = worksheet.max_column
                
                # Check if sheet has data
                has_data = max_row > 0 and max_col > 0
                
                # Try to detect if it looks like a table (has headers)
                has_headers = False
                header_row = None
                
                if has_data and max_row > 1:
                    # Check first few rows for potential headers
                    for row_num in range(1, min(4, max_row + 1)):
                        row_values = []
                        for col_num in range(1, min(max_col + 1, 11)):  # Check first 10 columns
                            cell_value = worksheet.cell(row=row_num, column=col_num).value
                            if cell_value is not None:
                                row_values.append(str(cell_value))
                        
                        # If row has multiple non-empty values, consider it potential headers
                        if len(row_values) >= 2:
                            has_headers = True
                            header_row = row_num
                            break
                
                # Get sample headers if found
                headers = []
                if has_headers and header_row:
                    for col_num in range(1, min(max_col + 1, 11)):  # First 10 columns
                        cell_value = worksheet.cell(row=header_row, column=col_num).value
                        if cell_value is not None:
                            headers.append(str(cell_value))
                        else:
                            headers.append(f"Column{col_num}")
                
                # Calculate approximate data rows
                data_rows = max(0, max_row - (header_row if header_row else 0))
                
                table_info = {
                    "name": sheet_name,
                    "type": "worksheet",
                    "has_data": has_data,
                    "has_headers": has_headers,
                    "header_row": header_row,
                    "total_rows": max_row,
                    "total_columns": max_col,
                    "data_rows": data_rows,
                    "sample_headers": headers[:5],  # First 5 headers
                    "description": self._generate_table_description(sheet_name, has_data, data_rows, max_col, headers)
                }
                
                tables.append(table_info)
            
            # Also check for named tables (Excel Table objects)
            for sheet_name in self.current_workbook.sheetnames:
                worksheet = self.current_workbook[sheet_name]
                
                # Check for Excel Table objects
                if hasattr(worksheet, 'tables') and worksheet.tables:
                    for table_name, table in worksheet.tables.items():
                        # Get table range and headers
                        table_range = table.ref
                        
                        # Parse table range to get dimensions
                        try:
                            from openpyxl.utils import range_boundaries
                            min_col, min_row, max_col, max_row = range_boundaries(table_range)
                            
                            table_headers = []
                            if table.tableStyleInfo and hasattr(table.tableStyleInfo, 'showFirstColumn'):
                                # Try to get actual headers
                                for col_num in range(min_col, max_col + 1):
                                    cell_value = worksheet.cell(row=min_row, column=col_num).value
                                    if cell_value:
                                        table_headers.append(str(cell_value))
                            
                            table_info = {
                                "name": table_name,
                                "type": "table",
                                "worksheet": sheet_name,
                                "has_data": True,
                                "has_headers": True,
                                "header_row": min_row,
                                "total_rows": max_row - min_row + 1,
                                "total_columns": max_col - min_col + 1,
                                "data_rows": max_row - min_row,
                                "sample_headers": table_headers[:5],
                                "range": table_range,
                                "description": f"Excel Table '{table_name}' in sheet '{sheet_name}' ({max_row - min_row} data rows)"
                            }
                            
                            tables.append(table_info)
                            
                        except Exception as e:
                            logger.warning(f"Error processing table {table_name}: {e}")
            
            logger.info(f"Found {len(tables)} tables/worksheets in {self.current_file_info['name']}")
            return tables
            
        except Exception as e:
            logger.error(f"Error extracting tables from Excel file: {e}")
            return []
    
    def _generate_table_description(self, sheet_name: str, has_data: bool, data_rows: int, total_cols: int, headers: List[str]) -> str:
        """Generate a human-readable description of the table/worksheet"""
        if not has_data:
            return f"Empty worksheet '{sheet_name}'"
        
        description = f"Worksheet '{sheet_name}'"
        
        if data_rows > 0:
            description += f" ({data_rows} rows, {total_cols} columns)"
            
            if headers:
                header_preview = ", ".join(headers[:3])
                if len(headers) > 3:
                    header_preview += "..."
                description += f" - Headers: {header_preview}"
        else:
            description += " (no data detected)"
        
        return description
    
    def get_file_info(self) -> Optional[Dict]:
        """Get information about the currently opened file"""
        return self.current_file_info
    
    def is_file_open(self) -> bool:
        """Check if a file is currently open"""
        return self.current_workbook is not None